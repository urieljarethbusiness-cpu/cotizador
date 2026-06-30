"""Excel Generator for Uriel Jareth Consulting — port of src/lib/excel-builder.ts using openpyxl.

Mantiene paridad de layout con el builder de TypeScript: mismas columnas, merges,
celdas dinámicas (wrapText + altura calculada) para que el texto nunca se salga.
"""

from __future__ import annotations

import io
import math
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.calculators import bucefalo_precio, detalle_modelo, calcular_totales_opcion, format_currency, precio_display, nota_demanda

FASES_MAP = {0: "FASE 0", 1: "FASE 1", 2: "FASE 2", 3: "FASE 3"}


def _argb(hex_color: str, alpha: str = "FF") -> str:
    h = hex_color.lstrip("#")
    if len(h) > 6:
        h = h[:6]
    return f"{alpha}{h.upper()}"


def _thin_border(color: str = "FFD1D5DB") -> Border:
    side = Side(style="thin", color=color)
    return Border(top=side, left=side, bottom=side, right=side)


def _sanitize_sheet_name(name: str) -> str:
    name = (name or "Servicio")[:31]
    return re.sub(r"[/\\*?\[\]:]", "", name) or "Servicio"


def _set_col_widths(ws: Worksheet, widths: list[tuple[str, float]]):
    for col_letter, width in widths:
        ws.column_dimensions[col_letter].width = width


def _merged_width(widths: dict[str, float], start_col: str, end_col: str) -> float:
    """Suma de anchos (unidades de carácter de Excel) entre dos columnas, inclusive."""
    start = column_index_from_string(start_col)
    end = column_index_from_string(end_col)
    return sum(widths.get(get_column_letter(c), 8) for c in range(start, end + 1))


def _estimate_row_height(text: str, width_chars: float, font_size: int = 10) -> float:
    """Altura (pt) que necesita una fila para mostrar `text` envuelto en `width_chars`."""
    cpl = max(8, int(width_chars))
    if font_size <= 10:
        line_h = 14.0
    elif font_size <= 11:
        line_h = 15.5
    elif font_size <= 12:
        line_h = 17.0
    else:
        line_h = 20.0
    lines = 0
    for seg in str(text or "").split("\n"):
        lines += max(1, math.ceil(len(seg) / cpl))
    return max(16.0, round(lines * line_h + 4))


def _set_wrapped(
    ws: Worksheet,
    row: int,
    col: str,
    text: str,
    width_chars: float,
    font_size: int = 10,
    horizontal: str = "left",
    vertical: str = "top",
):
    """Marca la celda como envolvente y agranda la fila si el texto lo requiere."""
    cell = ws[f"{col}{row}"]
    cell.alignment = Alignment(wrap_text=True, vertical=vertical, horizontal=horizontal)
    needed = _estimate_row_height(text, width_chars, font_size)
    current = ws.row_dimensions[row].height or 0
    if needed > current:
        ws.row_dimensions[row].height = needed


def generate_cotizacion_excel(data: dict[str, Any], saved: bool = False) -> bytes:
    """Generate a multi-sheet Excel workbook for a quotation.

    Args:
        data: dict with quotation data (same as PDF generator)
        saved: if True, uses real quotation number instead of "BORRADOR"
    """
    config_color = data.get("colorPrimario", "#2563eb")
    PRIMARY = _argb(config_color)
    PRIMARY_LIGHT = _argb(config_color, alpha="18")
    SECONDARY = _argb(data.get("colorSecundario", "#1e293b"))
    WHITE = "FFFFFFFF"
    LIGHT_BG = "FFF8FAFC"
    MUTED = "FF64748B"
    BORDER_COLOR = "FFE2E8F0"

    header_font = Font(bold=True, size=10, name="Calibri", color=WHITE)
    label_font = Font(bold=True, size=10, name="Calibri", color=MUTED)
    value_font = Font(size=10, name="Calibri")
    bold_font = Font(bold=True, size=10, name="Calibri")
    total_font = Font(bold=True, size=11, name="Calibri")
    small_font = Font(italic=True, size=9, name="Calibri", color=MUTED)
    fase_font = Font(bold=True, size=11, name="Calibri", color=PRIMARY)
    iva_font = Font(italic=False, size=9, name="Calibri", color=MUTED)
    beneficio_font = Font(bold=True, size=10, name="Calibri", color=PRIMARY)
    demanda_note_font = Font(italic=True, size=9, name="Calibri", color=PRIMARY)

    primary_fill = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
    primary_light_fill = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type="solid")
    secondary_fill = PatternFill(start_color=SECONDARY, end_color=SECONDARY, fill_type="solid")
    light_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    razon_social = data.get("razonSocial", "Uriel Jareth Consulting")
    domicilio_fiscal = data.get("domicilioFiscal")
    cliente_nombre = data.get("clienteNombre", "")
    cliente_empresa = data.get("clienteEmpresa", "")
    cliente_rfc = data.get("clienteRfc", "")
    asesor_nombre = data.get("asesorNombre", "")
    fecha = data.get("fecha", datetime.now())
    vigencia = data.get("vigencia", datetime.now())
    moneda = data.get("moneda", "MXN")
    tipo_cambio = data.get("tipoCambio", "NA")
    proyecto = data.get("proyecto", "MKT Digital")
    esquema = data.get("esquemaPago", "Pago Unico/Mensual")
    servicios = data.get("servicios", [])
    plan_nivel = data.get("planBucefaloNivel")
    plan_precio = data.get("planBucefaloPrecio")
    numero_label = data.get("numero", "BORRADOR") if saved else "BORRADOR"
    # IVA: por defecto se aplica (se factura). False = proyecto sin factura, precios finales.
    iva = data.get("incluirIva", True)
    iva_factor = 1.16 if iva else 1
    iva_suf = " + IVA" if iva else ""
    iva_total_lbl = " (c/IVA)" if iva else ""

    wb = Workbook()
    wb.properties.creator = "Uriel Jareth Consulting"

    # ═══════════════════════════════════════════════════════════════════════════
    # HOJA RESUMEN
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "HOJA RESUMEN"
    ws.sheet_view.showGridLines = False

    resumen_widths = {
        "A": 3, "B": 18, "C": 26, "D": 13, "E": 16,
        "F": 22, "G": 6, "H": 6, "I": 6, "J": 6, "K": 16,
    }
    _set_col_widths(ws, list(resumen_widths.items()))

    # Banner
    ws.merge_cells("B2:K2")
    cell = ws["B2"]
    cell.value = razon_social
    cell.font = Font(bold=True, size=16, name="Calibri", color=WHITE)
    cell.alignment = Alignment(vertical="center")
    for col in range(1, 12):
        ws.cell(row=2, column=col).fill = primary_fill
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:K3")
    cell = ws["B3"]
    cell.value = numero_label
    cell.font = Font(bold=True, size=11, name="Calibri", color=WHITE)
    cell.alignment = Alignment(vertical="center")
    for col in range(1, 12):
        ws.cell(row=3, column=col).fill = secondary_fill

    # Info section
    info_start = 5
    info_rows = [
        [("B", "En atencion a:"), ("C", cliente_nombre), ("E", "Empresa:"), ("F", cliente_empresa or "—")],
        [("B", "Proyecto:"), ("C", proyecto), ("E", "Moneda:"), ("F", moneda)],
        [("B", "Asesor:"), ("C", asesor_nombre), ("E", "Tipo de Cambio:"), ("F", tipo_cambio)],
        [("B", "Fecha:"), ("C", ""), ("E", "Esquema de Pago:"), ("F", esquema)],
    ]
    c_val_w = _merged_width(resumen_widths, "C", "C")
    f_val_w = _merged_width(resumen_widths, "F", "F")

    for i, row_data in enumerate(info_rows):
        r = info_start + i
        (lc1, lv1), (vc1, vv1), (lc2, lv2), (vc2, vv2) = row_data
        ws[f"{lc1}{r}"].value = lv1
        ws[f"{lc1}{r}"].font = label_font
        ws[f"{vc1}{r}"].value = vv1
        ws[f"{vc1}{r}"].font = value_font
        if vv1:
            _set_wrapped(ws, r, vc1, vv1, c_val_w)
        ws[f"{lc2}{r}"].value = lv2
        ws[f"{lc2}{r}"].font = label_font
        ws[f"{vc2}{r}"].value = vv2
        ws[f"{vc2}{r}"].font = value_font
        if vv2:
            _set_wrapped(ws, r, vc2, vv2, f_val_w)

    # Fecha (en la fila "Fecha:")
    fecha_cell = ws[f"C{info_start + 3}"]
    fecha_cell.value = fecha
    fecha_cell.number_format = "DD/MM/YYYY"
    fecha_cell.font = value_font

    vig_row = info_start + 4
    ws[f"B{vig_row}"].value = "Vigencia:"
    ws[f"B{vig_row}"].font = label_font
    ws[f"C{vig_row}"].value = vigencia
    ws[f"C{vig_row}"].number_format = "DD/MM/YYYY"
    ws[f"C{vig_row}"].font = value_font
    # RFC del cliente: solo se muestra si fue capturado (proyectos con factura).
    if cliente_rfc:
        ws[f"E{vig_row}"].value = "RFC:"
        ws[f"E{vig_row}"].font = label_font
        ws[f"F{vig_row}"].value = cliente_rfc
        ws[f"F{vig_row}"].font = value_font

    # Services table
    table_start = info_start + 6
    col_b = column_index_from_string("B")
    col_k = column_index_from_string("K")

    def style_table_row(row_num: int, fill: PatternFill, border_color: str):
        """Estiliza (relleno + borde) todo el ancho de la fila, de B a K."""
        border = _thin_border(border_color)
        for c in range(col_b, col_k + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = fill
            cell.border = border

    ws.merge_cells(f"D{table_start}:J{table_start}")
    style_table_row(table_start, primary_fill, PRIMARY)
    for col_letter, txt, align in [
        ("B", "Fase", "left"),
        ("C", "Tipo de Pago", "left"),
        ("D", "Servicio", "left"),
        ("K", "Precio", "right"),
    ]:
        cell = ws[f"{col_letter}{table_start}"]
        cell.value = txt
        cell.font = header_font
        cell.alignment = Alignment(horizontal=align, vertical="center")
    ws.row_dimensions[table_start].height = 24

    # Las partidas "demanda" salen de la tabla normal y se muestran en su propio modulo.
    demanda_servs = [s for s in servicios if s.get("modeloCobro") == "demanda"]
    comprometidos = [s for s in servicios if s.get("modeloCobro") != "demanda"]
    servicios_unicos = [s for s in comprometidos if s.get("tipoPago") == "unico"]
    servicios_mensuales = [s for s in comprometidos if s.get("tipoPago") == "mensual"]
    serv_width = _merged_width(resumen_widths, "D", "J")
    es_doble = bool(data.get("esDoble"))
    opciones_meta = data.get("opcionesMetadata") or {}

    row = table_start + 1

    # Totales (helpers reutilizados por modo normal y doble propuesta)
    def write_total_row(label: str, amount: float, font: Font):
        nonlocal row
        ws.merge_cells(f"B{row}:J{row}")
        ws[f"B{row}"].value = label
        ws[f"B{row}"].font = font
        ws[f"B{row}"].alignment = Alignment(vertical="center")
        ws[f"K{row}"].value = amount
        ws[f"K{row}"].number_format = "$#,##0.00"
        ws[f"K{row}"].font = font
        ws[f"K{row}"].alignment = Alignment(horizontal="right", vertical="center")
        style_table_row(row, primary_light_fill, PRIMARY)
        row += 1

    def write_iva_note():
        nonlocal row
        if not iva:  # proyecto sin factura: no se anota "+ IVA"
            return
        ws.merge_cells(f"B{row}:J{row}")
        ws[f"B{row}"].value = "(+ IVA)"
        ws[f"B{row}"].font = iva_font
        row += 1

    def draw_esquema_horas(servs, titulo_opcion=None):
        """Modulo dedicado del esquema por horas: tarjeta de tarifas + beneficios + nota."""
        nonlocal row
        if not servs:
            return
        b_idx = column_index_from_string("B")
        k_idx = column_index_from_string("K")
        # Encabezado del modulo
        ws.merge_cells(f"B{row}:K{row}")
        ws[f"B{row}"].value = "ESQUEMA DE TRABAJO POR HORAS" + (f" - {titulo_opcion}" if titulo_opcion else "")
        ws[f"B{row}"].font = Font(bold=True, size=12, name="Calibri", color=WHITE)
        ws[f"B{row}"].alignment = Alignment(vertical="center")
        for col in range(b_idx, k_idx + 1):
            ws.cell(row=row, column=col).fill = primary_fill
        ws.row_dimensions[row].height = 22
        row += 1

        for serv in servs:
            ws.merge_cells(f"B{row}:H{row}")
            ws[f"B{row}"].value = serv.get("nombre", "")
            ws[f"B{row}"].font = bold_font
            ws[f"B{row}"].alignment = Alignment(vertical="center")
            ws.merge_cells(f"I{row}:K{row}")
            ws[f"I{row}"].value = f"{precio_display(serv)}{iva_suf}"
            ws[f"I{row}"].font = Font(bold=True, size=10, name="Calibri", color=PRIMARY)
            ws[f"I{row}"].alignment = Alignment(horizontal="right", vertical="center")
            for col in range(b_idx, k_idx + 1):
                ws.cell(row=row, column=col).border = _thin_border(BORDER_COLOR)
            row += 1
            for e in serv.get("entregables", []):
                ws.merge_cells(f"B{row}:K{row}")
                txt = f"• {e}"
                ws[f"B{row}"].value = txt
                ws[f"B{row}"].font = value_font
                _set_wrapped(ws, row, "B", txt, _merged_width(resumen_widths, "B", "K"))
                row += 1
            for b in serv.get("beneficios", []):
                ws.merge_cells(f"B{row}:K{row}")
                txt = f"» {b}"
                ws[f"B{row}"].value = txt
                ws[f"B{row}"].font = beneficio_font
                _set_wrapped(ws, row, "B", txt, _merged_width(resumen_widths, "B", "K"))
                row += 1

        # Total tipo "Segun consumo"
        ws.merge_cells(f"B{row}:J{row}")
        ws[f"B{row}"].value = "Total mensual"
        ws[f"B{row}"].font = total_font
        ws[f"B{row}"].alignment = Alignment(vertical="center")
        ws[f"K{row}"].value = "Segun consumo"
        ws[f"K{row}"].font = total_font
        ws[f"K{row}"].alignment = Alignment(horizontal="right", vertical="center")
        style_table_row(row, primary_light_fill, PRIMARY)
        row += 1

        nota = "Facturacion a fin de mes segun las horas efectivamente consumidas. No requiere anticipo. Precios en MXN" + (", no incluyen IVA." if iva else ", no se emite factura.")
        ws.merge_cells(f"B{row}:K{row}")
        ws[f"B{row}"].value = nota
        ws[f"B{row}"].font = small_font
        _set_wrapped(ws, row, "B", nota, _merged_width(resumen_widths, "B", "K"), font_size=9)
        row += 2

    def draw_service_rows(servs):
        nonlocal row
        current_fase = -1
        for serv in servs:
            fase = serv.get("fase", 0)
            if fase != current_fase:
                current_fase = fase
                ws.merge_cells(f"B{row}:K{row}")
                ws[f"B{row}"].value = FASES_MAP.get(fase, f"FASE {fase}")
                ws[f"B{row}"].font = fase_font
                ws[f"B{row}"].alignment = Alignment(vertical="center")
                for col in range(2, 12):
                    ws.cell(row=row, column=col).fill = primary_light_fill
                ws.row_dimensions[row].height = 20
                row += 1

            row_bg = light_fill if row % 2 == 0 else white_fill
            ws.merge_cells(f"D{row}:J{row}")
            style_table_row(row, row_bg, BORDER_COLOR)
            ws[f"C{row}"].value = "Unico" if serv.get("tipoPago") == "unico" else "Mensual"
            ws[f"C{row}"].font = value_font
            ws[f"C{row}"].alignment = Alignment(vertical="top")
            _detalle = detalle_modelo(serv)
            serv_text = f"{serv.get('nombre', '')}  —  {_detalle}" if _detalle else serv.get("nombre", "")
            ws[f"D{row}"].value = serv_text
            ws[f"D{row}"].font = value_font
            _set_wrapped(ws, row, "D", serv_text, serv_width)
            # "demanda": muestra la tarifa/hr (texto) en vez de $0; nunca suma al total.
            if serv.get("modeloCobro") == "demanda":
                ws[f"K{row}"].value = precio_display(serv)
                ws[f"K{row}"].font = value_font
                ws[f"K{row}"].alignment = Alignment(horizontal="right", vertical="top")
            else:
                ws[f"K{row}"].value = serv.get("precio", 0)
                ws[f"K{row}"].number_format = "$#,##0.00"
                ws[f"K{row}"].font = value_font
                ws[f"K{row}"].alignment = Alignment(horizontal="right", vertical="top")
            row += 1

            # Beneficios destacados (denotan valor; util en partidas demanda sin precio).
            for b in serv.get("beneficios", []):
                ws.merge_cells(f"D{row}:J{row}")
                txt = f"» {b}"
                ws[f"D{row}"].value = txt
                ws[f"D{row}"].font = beneficio_font
                _set_wrapped(ws, row, "D", txt, serv_width)
                row += 1

    def write_opcion_banner(op: str):
        nonlocal row
        meta = opciones_meta.get(op) or {}
        ws.merge_cells(f"B{row}:K{row}")
        titulo = meta.get("titulo")
        ws[f"B{row}"].value = f"OPCION {op}" + (f": {titulo}" if titulo else "")
        ws[f"B{row}"].font = Font(bold=True, size=12, name="Calibri", color=WHITE)
        ws[f"B{row}"].alignment = Alignment(vertical="center")
        for col in range(2, 12):
            ws.cell(row=row, column=col).fill = secondary_fill
        ws.row_dimensions[row].height = 22
        row += 1
        desc = meta.get("descripcion")
        if desc:
            ws.merge_cells(f"B{row}:K{row}")
            ws[f"B{row}"].value = desc
            ws[f"B{row}"].font = value_font
            _set_wrapped(ws, row, "B", desc, _merged_width(resumen_widths, "B", "K"))
            row += 1
        no_incluye = meta.get("noIncluye")
        if no_incluye:
            ws.merge_cells(f"B{row}:K{row}")
            ws[f"B{row}"].value = f"No incluye: {no_incluye}"
            ws[f"B{row}"].font = small_font
            _set_wrapped(ws, row, "B", f"No incluye: {no_incluye}", _merged_width(resumen_widths, "B", "K"), font_size=9)
            row += 1

    if es_doble:
        for op in ("1", "2"):
            u = [s for s in servicios_unicos if s.get("opcion") in (op, "ambas")]
            me = [s for s in servicios_mensuales if s.get("opcion") in (op, "ambas")]
            dem = [s for s in demanda_servs if s.get("opcion") in (op, "ambas")]
            write_opcion_banner(op)
            if u or me:
                draw_service_rows(u + me)
                row += 1
                write_total_row(f"Total Pago Unico - Opcion {op}", sum(s.get("precio", 0) for s in u), total_font)
                write_iva_note()
                write_total_row(f"Total Pago Mensual - Opcion {op}", sum(s.get("precio", 0) for s in me), total_font)
                write_iva_note()
            draw_esquema_horas(dem)
            row += 1

        # Comparativa de opciones (totales con IVA + horas)
        t1 = calcular_totales_opcion(servicios, "1")
        t2 = calcular_totales_opcion(servicios, "2")
        ws.merge_cells(f"B{row}:K{row}")
        ws[f"B{row}"].value = "COMPARATIVA DE OPCIONES"
        ws[f"B{row}"].font = Font(bold=True, size=12, name="Calibri", color=WHITE)
        ws[f"B{row}"].alignment = Alignment(vertical="center")
        for col in range(2, 12):
            ws.cell(row=row, column=col).fill = primary_fill
        ws.row_dimensions[row].height = 22
        row += 1
        t1_tit = (opciones_meta.get("1") or {}).get("titulo")
        t2_tit = (opciones_meta.get("2") or {}).get("titulo")

        def comp_row(lab: str, v1: str, v2: str, font: Font):
            nonlocal row
            ws.merge_cells(f"B{row}:F{row}")
            ws[f"B{row}"].value = lab
            ws[f"B{row}"].font = font
            ws.merge_cells(f"G{row}:H{row}")
            ws[f"G{row}"].value = v1
            ws[f"G{row}"].font = font
            ws[f"G{row}"].alignment = Alignment(horizontal="right")
            ws.merge_cells(f"I{row}:K{row}")
            ws[f"I{row}"].value = v2
            ws[f"I{row}"].font = font
            ws[f"I{row}"].alignment = Alignment(horizontal="right")
            border = _thin_border(BORDER_COLOR)
            for col in range(col_b, col_k + 1):
                ws.cell(row=row, column=col).border = border
            row += 1

        op1_dem = any(s.get("modeloCobro") == "demanda" and s.get("opcion") in ("1", "ambas") for s in servicios)
        op2_dem = any(s.get("modeloCobro") == "demanda" and s.get("opcion") in ("2", "ambas") for s in servicios)
        mens1 = "Segun consumo" if t1["totalMensual"] == 0 and op1_dem else format_currency(t1["totalMensual"] * iva_factor)
        mens2 = "Segun consumo" if t2["totalMensual"] == 0 and op2_dem else format_currency(t2["totalMensual"] * iva_factor)
        comp_row("Concepto", "Opcion 1" + (f" - {t1_tit}" if t1_tit else ""), "Opcion 2" + (f" - {t2_tit}" if t2_tit else ""), bold_font)
        comp_row(f"Total unico{iva_total_lbl}", format_currency(t1["totalUnico"] * iva_factor), format_currency(t2["totalUnico"] * iva_factor), value_font)
        comp_row(f"Total mensual{iva_total_lbl}", mens1, mens2, value_font)
        comp_row("Horas estimadas", f"{t1['horas']:g} h", f"{t2['horas']:g} h", value_font)
        row += 1
    else:
        if servicios_unicos or servicios_mensuales:
            draw_service_rows(servicios_unicos + servicios_mensuales)
            row += 1
            write_total_row("Total Pago Unico", sum(s.get("precio", 0) for s in servicios_unicos), total_font)
            write_iva_note()
            write_total_row("Total Pago Mensual", sum(s.get("precio", 0) for s in servicios_mensuales), total_font)
            write_iva_note()
        draw_esquema_horas(demanda_servs)
        row += 1

    if plan_nivel:
        precio = plan_precio if plan_precio is not None else bucefalo_precio(plan_nivel)
        write_total_row(f"CRM Bucefalo - {plan_nivel.capitalize()}", precio, bold_font)
        row += 1

    ws.merge_cells(f"B{row}:K{row}")
    nota_resumen = (
        "Todos los precios son en Moneda Nacional (MX), los precios no incluyen IVA. Vigencia de 15 dias habiles."
        if iva
        else "Todos los precios son en Moneda Nacional (MX). Proyecto sin factura (no se aplica IVA). Vigencia de 15 dias habiles."
    )
    ws[f"B{row}"].value = nota_resumen
    ws[f"B{row}"].font = small_font
    _set_wrapped(ws, row, "B", nota_resumen, _merged_width(resumen_widths, "B", "K"), font_size=9)

    # ═══════════════════════════════════════════════════════════════════════════
    # HOJAS DETALLADAS POR SERVICIO
    # ═══════════════════════════════════════════════════════════════════════════
    detail_widths = {
        "A": 3, "B": 16, "C": 30, "D": 8, "E": 4,
        "F": 18, "G": 22, "H": 6, "I": 6, "J": 16,
    }
    used_names: set[str] = set()
    for serv in servicios:
        base = _sanitize_sheet_name(serv.get("nombre", "Servicio"))
        safe_name = base
        n = 2
        while safe_name.lower() in used_names:
            suffix = f" ({n})"
            safe_name = base[: 31 - len(suffix)] + suffix
            n += 1
        used_names.add(safe_name.lower())

        dw = wb.create_sheet(title=safe_name)
        dw.sheet_view.showGridLines = False
        _set_col_widths(dw, list(detail_widths.items()))

        fase = serv.get("fase", 0)
        dw.merge_cells("B1:J1")
        banner_text = f"{FASES_MAP.get(fase, f'FASE {fase}')} — {serv.get('nombre', '')}"
        dw["B1"].value = banner_text
        dw["B1"].font = Font(bold=True, size=14, name="Calibri", color=WHITE)
        dw["B1"].alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, 11):
            dw.cell(row=1, column=col).fill = primary_fill
        dw.row_dimensions[1].height = max(
            30, _estimate_row_height(banner_text, _merged_width(detail_widths, "B", "J"), 14)
        )

        c_val = _merged_width(detail_widths, "C", "D")
        g_val = _merged_width(detail_widths, "G", "H")

        r = 3
        dw[f"B{r}"].value = "Cliente:"
        dw[f"B{r}"].font = label_font
        dw.merge_cells(f"C{r}:D{r}")
        dw[f"C{r}"].value = cliente_empresa or cliente_nombre
        dw[f"C{r}"].font = bold_font
        _set_wrapped(dw, r, "C", cliente_empresa or cliente_nombre, c_val)
        dw[f"F{r}"].value = "No. Cotización:"
        dw[f"F{r}"].font = label_font
        dw.merge_cells(f"G{r}:H{r}")
        dw[f"G{r}"].value = numero_label
        dw[f"G{r}"].font = bold_font
        r += 2

        dw[f"B{r}"].value = "Servicio:"
        dw[f"B{r}"].font = label_font
        dw.merge_cells(f"C{r}:D{r}")
        dw[f"C{r}"].value = serv.get("nombre", "")
        dw[f"C{r}"].font = Font(bold=True, size=12, name="Calibri", color=PRIMARY)
        _set_wrapped(dw, r, "C", serv.get("nombre", ""), c_val, font_size=12)
        dw[f"F{r}"].value = "Tiempo de Entrega:"
        dw[f"F{r}"].font = label_font
        dw.merge_cells(f"G{r}:H{r}")
        dw[f"G{r}"].value = serv.get("tiempoEntrega", "")
        dw[f"G{r}"].font = value_font
        _set_wrapped(dw, r, "G", serv.get("tiempoEntrega", ""), g_val)
        r += 2

        _detalle_serv = detalle_modelo(serv)
        if _detalle_serv:
            dw[f"B{r}"].value = "Detalle:"
            dw[f"B{r}"].font = label_font
            dw.merge_cells(f"C{r}:J{r}")
            dw[f"C{r}"].value = _detalle_serv
            dw[f"C{r}"].font = value_font
            _set_wrapped(dw, r, "C", _detalle_serv, _merged_width(detail_widths, "C", "J"))
            r += 2

        # Encabezado de entregables — barra completa B:J
        dw.merge_cells(f"B{r}:J{r}")
        dw[f"B{r}"].value = "Entregables"
        dw[f"B{r}"].font = header_font
        dw[f"B{r}"].alignment = Alignment(vertical="center")
        for col in range(column_index_from_string("B"), column_index_from_string("J") + 1):
            dw.cell(row=r, column=col).fill = primary_fill
        dw.row_dimensions[r].height = 22
        r += 1

        ent_width = _merged_width(detail_widths, "B", "J")
        entregables = serv.get("entregables", [])
        b_idx = column_index_from_string("B")
        j_idx = column_index_from_string("J")
        for i, ent in enumerate(entregables):
            er = r + i
            dw.merge_cells(f"B{er}:J{er}")
            txt = f"{i + 1}. {ent}"
            dw[f"B{er}"].value = txt
            dw[f"B{er}"].font = value_font
            _set_wrapped(dw, er, "B", txt, ent_width)
            bg = light_fill if i % 2 == 1 else white_fill
            for col in range(b_idx, j_idx + 1):
                dw.cell(row=er, column=col).fill = bg
                dw.cell(row=er, column=col).border = _thin_border(BORDER_COLOR)
        r += len(entregables) + 1

        # Beneficios (si los hay): seccion destacada en color primario.
        beneficios = serv.get("beneficios", [])
        if beneficios:
            dw.merge_cells(f"B{r}:J{r}")
            dw[f"B{r}"].value = "Beneficios"
            dw[f"B{r}"].font = header_font
            dw[f"B{r}"].alignment = Alignment(vertical="center")
            for col in range(b_idx, j_idx + 1):
                dw.cell(row=r, column=col).fill = primary_fill
            dw.row_dimensions[r].height = 22
            r += 1
            for i, ben in enumerate(beneficios):
                er = r + i
                dw.merge_cells(f"B{er}:J{er}")
                txt = f"» {ben}"
                dw[f"B{er}"].value = txt
                dw[f"B{er}"].font = beneficio_font
                _set_wrapped(dw, er, "B", txt, ent_width)
                for col in range(b_idx, j_idx + 1):
                    dw.cell(row=er, column=col).border = _thin_border(BORDER_COLOR)
            r += len(beneficios) + 1

        # Total: etiqueta B:H + precio I:J
        dw.merge_cells(f"B{r}:H{r}")
        tipo_label = "Unico" if serv.get("tipoPago") == "unico" else "Mensual"
        dw[f"B{r}"].value = f"Total Pago {tipo_label}"
        dw[f"B{r}"].font = total_font
        dw[f"B{r}"].alignment = Alignment(vertical="center")
        dw.merge_cells(f"I{r}:J{r}")
        if serv.get("modeloCobro") == "demanda":
            dw[f"I{r}"].value = precio_display(serv)
        else:
            dw[f"I{r}"].value = serv.get("precio", 0)
            dw[f"I{r}"].number_format = "$#,##0.00"
        dw[f"I{r}"].font = total_font
        dw[f"I{r}"].alignment = Alignment(horizontal="right", vertical="center")
        for col in range(b_idx, j_idx + 1):
            dw.cell(row=r, column=col).fill = primary_light_fill
            dw.cell(row=r, column=col).border = _thin_border(PRIMARY)
        r += 1
        if serv.get("modeloCobro") == "demanda":
            dw[f"B{r}"].value = f"(Se factura a fin de mes segun consumo{iva_suf})"
        else:
            dw[f"B{r}"].value = "(+ IVA)" if iva else "(Sin factura)"
        dw[f"B{r}"].font = iva_font
        r += 2

        dw.merge_cells(f"B{r}:J{r}")
        nota_detalle = "Esta cotización tiene una vigencia de 15 dias habiles. Todos los precios en MXN" + (", precios + IVA." if iva else ", sin factura.")
        dw[f"B{r}"].value = nota_detalle
        dw[f"B{r}"].font = small_font
        _set_wrapped(dw, r, "B", nota_detalle, ent_width, font_size=9)
        r += 2

        dw.merge_cells(f"B{r}:J{r}")
        dw[f"B{r}"].value = razon_social
        dw[f"B{r}"].font = Font(bold=True, size=10, name="Calibri", color=PRIMARY)
        if domicilio_fiscal:
            r += 1
            dw.merge_cells(f"B{r}:J{r}")
            dw[f"B{r}"].value = domicilio_fiscal
            dw[f"B{r}"].font = iva_font
            _set_wrapped(dw, r, "B", domicilio_fiscal, ent_width, font_size=9)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
